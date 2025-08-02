import pandas as pd
import numpy as np
from typing import Dict, List, Any, Optional, Union
import re
from datetime import datetime, date
import dateutil.parser


def validate_amount_format(value: str) -> Dict[str, Any]:
    """
    Validate amount format and return parsing result.
    
    Args:
        value: String value to validate
        
    Returns:
        Dictionary with validation results
    """
    if not value or pd.isna(value):
        return {
            'is_valid': False,
            'error': 'Empty or null value',
            'parsed_value': None
        }
    
    # Common amount patterns
    patterns = {
        'us_currency': r'^\$[\d,]+\.?\d*$',
        'european_currency': r'^€[\d.,]+$',
        'indian_currency': r'^₹[\d,]+\.?\d*$',
        'negative_parentheses': r'^\([\d,]+\.?\d*\)$',
        'trailing_negative': r'^[\d,]+\.?\d*-$',
        'abbreviated': r'^[\d.]+[KMBT]$',
        'plain_number': r'^[\d,]+\.?\d*$'
    }
    
    str_value = str(value).strip()
    
    for pattern_name, pattern in patterns.items():
        if re.match(pattern, str_value, re.IGNORECASE):
            return {
                'is_valid': True,
                'pattern': pattern_name,
                'parsed_value': str_value
            }
    
    return {
        'is_valid': False,
        'error': f'Invalid amount format: {str_value}',
        'parsed_value': None
    }


def validate_date_format(value: str) -> Dict[str, Any]:
    """
    Validate date format and return parsing result.
    
    Args:
        value: String value to validate
        
    Returns:
        Dictionary with validation results
    """
    if not value or pd.isna(value):
        return {
            'is_valid': False,
            'error': 'Empty or null value',
            'parsed_value': None
        }
    
    # Common date patterns
    patterns = {
        'mm_dd_yyyy': r'^\d{1,2}/\d{1,2}/\d{4}$',
        'dd_mm_yyyy': r'^\d{1,2}/\d{1,2}/\d{4}$',
        'yyyy_mm_dd': r'^\d{4}-\d{1,2}-\d{1,2}$',
        'dd_mon_yyyy': r'^\d{1,2}-\w{3}-\d{4}$',
        'quarter': r'^Q[1-4]\s+\d{4}$',
        'quarter_short': r'^Q[1-4]-\d{2}$',
        'month_year': r'^\w+\s+\d{4}$',
        'excel_serial': r'^\d{5}$'  # Excel serial dates are typically 5 digits
    }
    
    str_value = str(value).strip()
    
    # Check for Excel serial date first
    try:
        float_val = float(str_value)
        if 1 <= float_val <= 100000:  # Reasonable Excel date range
            return {
                'is_valid': True,
                'pattern': 'excel_serial',
                'parsed_value': float_val
            }
    except ValueError:
        pass
    
    # Check other patterns
    for pattern_name, pattern in patterns.items():
        if re.match(pattern, str_value, re.IGNORECASE):
            return {
                'is_valid': True,
                'pattern': pattern_name,
                'parsed_value': str_value
            }
    
    # Try dateutil parser as fallback
    try:
        parsed_date = dateutil.parser.parse(str_value)
        return {
            'is_valid': True,
            'pattern': 'dateutil_parsed',
            'parsed_value': parsed_date
        }
    except:
        pass
    
    return {
        'is_valid': False,
        'error': f'Invalid date format: {str_value}',
        'parsed_value': None
    }


def validate_column_data_type(series: pd.Series, expected_type: str) -> Dict[str, Any]:
    """
    Validate if a column matches the expected data type.
    
    Args:
        series: Pandas series to validate
        expected_type: Expected data type ('string', 'number', 'date')
        
    Returns:
        Dictionary with validation results
    """
    if series.empty:
        return {
            'is_valid': False,
            'error': 'Empty series',
            'match_percentage': 0.0
        }
    
    # Remove null values for analysis
    clean_series = series.dropna()
    
    if len(clean_series) == 0:
        return {
            'is_valid': False,
            'error': 'Series contains only null values',
            'match_percentage': 0.0
        }
    
    valid_count = 0
    total_count = len(clean_series)
    
    if expected_type == 'number':
        # Check if values can be converted to numeric
        numeric_series = pd.to_numeric(clean_series, errors='coerce')
        valid_count = numeric_series.notna().sum()
    
    elif expected_type == 'date':
        # Check if values can be parsed as dates
        for value in clean_series:
            if validate_date_format(str(value))['is_valid']:
                valid_count += 1
    
    elif expected_type == 'string':
        # String type is always valid (everything can be string)
        valid_count = total_count
    
    match_percentage = (valid_count / total_count) * 100
    
    return {
        'is_valid': match_percentage >= 80,  # 80% threshold
        'match_percentage': round(match_percentage, 2),
        'valid_count': valid_count,
        'total_count': total_count,
        'expected_type': expected_type
    }


def validate_dataframe_schema(df: pd.DataFrame, schema: Dict[str, str]) -> Dict[str, Any]:
    """
    Validate DataFrame against a schema definition.
    
    Args:
        df: DataFrame to validate
        schema: Dictionary mapping column names to expected types
        
    Returns:
        Dictionary with validation results
    """
    validation_result = {
        'is_valid': True,
        'errors': [],
        'warnings': [],
        'column_validations': {}
    }
    
    # Check if all expected columns exist
    missing_columns = set(schema.keys()) - set(df.columns)
    if missing_columns:
        validation_result['errors'].append(f'Missing columns: {list(missing_columns)}')
        validation_result['is_valid'] = False
    
    # Validate each column
    for column, expected_type in schema.items():
        if column not in df.columns:
            continue
        
        col_validation = validate_column_data_type(df[column], expected_type)
        validation_result['column_validations'][column] = col_validation
        
        if not col_validation['is_valid']:
            validation_result['errors'].append(
                f'Column "{column}": {col_validation.get("error", "Type mismatch")}'
            )
            validation_result['is_valid'] = False
        elif col_validation['match_percentage'] < 100:
            validation_result['warnings'].append(
                f'Column "{column}": {col_validation["match_percentage"]}% match with {expected_type}'
            )
    
    return validation_result


def validate_financial_data_quality(df: pd.DataFrame) -> Dict[str, Any]:
    """
    Validate financial data quality with specific checks.
    
    Args:
        df: DataFrame to validate
        
    Returns:
        Dictionary with quality validation results
    """
    quality_report = {
        'overall_score': 0,
        'checks_passed': 0,
        'total_checks': 0,
        'issues': [],
        'recommendations': []
    }
    
    checks = []
    
    # Check 1: No completely empty columns
    empty_columns = df.columns[df.isnull().all()].tolist()
    if empty_columns:
        quality_report['issues'].append(f'Empty columns found: {empty_columns}')
    else:
        quality_report['checks_passed'] += 1
    quality_report['total_checks'] += 1
    
    # Check 2: No duplicate rows
    duplicate_count = df.duplicated().sum()
    if duplicate_count > 0:
        quality_report['issues'].append(f'Found {duplicate_count} duplicate rows')
    else:
        quality_report['checks_passed'] += 1
    quality_report['total_checks'] += 1
    
    # Check 3: Reasonable null percentage (< 50%)
    null_percentages = (df.isnull().sum() / len(df)) * 100
    high_null_columns = null_percentages[null_percentages > 50].index.tolist()
    if high_null_columns:
        quality_report['issues'].append(f'High null percentage columns: {high_null_columns}')
    else:
        quality_report['checks_passed'] += 1
    quality_report['total_checks'] += 1
    
    # Check 4: Date columns are valid dates
    date_columns = []
    for col in df.columns:
        if any(keyword in col.lower() for keyword in ['date', 'time', 'created', 'updated']):
            date_columns.append(col)
    
    if date_columns:
        date_validation_passed = True
        for col in date_columns:
            col_validation = validate_column_data_type(df[col], 'date')
            if not col_validation['is_valid']:
                quality_report['issues'].append(f'Date column "{col}" validation failed')
                date_validation_passed = False
        
        if date_validation_passed:
            quality_report['checks_passed'] += 1
        quality_report['total_checks'] += 1
    
    # Check 5: Amount columns are numeric
    amount_columns = []
    for col in df.columns:
        if any(keyword in col.lower() for keyword in ['amount', 'value', 'price', 'cost', 'revenue', 'balance']):
            amount_columns.append(col)
    
    if amount_columns:
        amount_validation_passed = True
        for col in amount_columns:
            col_validation = validate_column_data_type(df[col], 'number')
            if not col_validation['is_valid']:
                quality_report['issues'].append(f'Amount column "{col}" validation failed')
                amount_validation_passed = False
        
        if amount_validation_passed:
            quality_report['checks_passed'] += 1
        quality_report['total_checks'] += 1
    
    # Calculate overall score
    if quality_report['total_checks'] > 0:
        quality_report['overall_score'] = round(
            (quality_report['checks_passed'] / quality_report['total_checks']) * 100, 2
        )
    
    # Generate recommendations
    if quality_report['overall_score'] < 80:
        quality_report['recommendations'].append('Data quality needs improvement')
    
    if duplicate_count > 0:
        quality_report['recommendations'].append('Remove duplicate rows')
    
    if high_null_columns:
        quality_report['recommendations'].append('Investigate high null percentage columns')
    
    return quality_report


def validate_file_format(file_path: str) -> Dict[str, Any]:
    """
    Validate file format and structure.
    
    Args:
        file_path: Path to the file
        
    Returns:
        Dictionary with validation results
    """
    import os
    
    validation_result = {
        'is_valid': False,
        'file_type': None,
        'error': None,
        'file_info': {}
    }
    
    # Check if file exists
    if not os.path.exists(file_path):
        validation_result['error'] = 'File does not exist'
        return validation_result
    
    # Get file extension
    file_ext = os.path.splitext(file_path)[1].lower()
    
    # Validate Excel files
    if file_ext in ['.xlsx', '.xls']:
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            validation_result.update({
                'is_valid': True,
                'file_type': 'excel',
                'sheets': workbook.sheetnames,
                'sheet_count': len(workbook.sheetnames)
            })
        except Exception as e:
            validation_result['error'] = f'Invalid Excel file: {str(e)}'
    
    # Validate CSV files
    elif file_ext == '.csv':
        try:
            df = pd.read_csv(file_path, nrows=5)  # Read first 5 rows to validate
            validation_result.update({
                'is_valid': True,
                'file_type': 'csv',
                'columns': list(df.columns),
                'sample_rows': len(df)
            })
        except Exception as e:
            validation_result['error'] = f'Invalid CSV file: {str(e)}'
    
    else:
        validation_result['error'] = f'Unsupported file format: {file_ext}'
    
    # Add file information
    if validation_result['is_valid']:
        stat = os.stat(file_path)
        validation_result['file_info'] = {
            'size_bytes': stat.st_size,
            'size_mb': round(stat.st_size / (1024 * 1024), 2),
            'modified_time': datetime.fromtimestamp(stat.st_mtime)
        }
    
    return validation_result 