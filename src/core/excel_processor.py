import pandas as pd
import openpyxl
from typing import Dict, List, Tuple, Optional
import os


class ExcelProcessor:
    """
    Handles Excel file reading and processing with support for multiple worksheets.
    """
    
    def __init__(self):
        self.loaded_files = {}
        self.file_info = {}
    
    def load_files(self, file_paths: List[str]) -> Dict[str, Dict]:
        """
        Load multiple Excel files and store their information.
        
        Args:
            file_paths: List of paths to Excel files
            
        Returns:
            Dictionary containing file information for each loaded file
        """
        for file_path in file_paths:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File not found: {file_path}")
            
            try:
                # Load with openpyxl for detailed sheet info
                workbook = openpyxl.load_workbook(file_path, read_only=True)
                
                # Load with pandas for data processing
                excel_file = pd.ExcelFile(file_path)
                
                file_info = {
                    'file_path': file_path,
                    'sheets': workbook.sheetnames,
                    'sheet_count': len(workbook.sheetnames),
                    'excel_file': excel_file,
                    'workbook': workbook
                }
                
                self.loaded_files[file_path] = file_info
                self.file_info[file_path] = file_info
                
            except Exception as e:
                raise Exception(f"Error loading file {file_path}: {str(e)}")
        
        return self.file_info
    
    def get_sheet_info(self, file_path: str) -> Dict:
        """
        Get detailed information about all sheets in a file.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Dictionary with sheet information
        """
        if file_path not in self.loaded_files:
            raise ValueError(f"File {file_path} not loaded. Use load_files() first.")
        
        file_info = self.loaded_files[file_path]
        sheet_info = {}
        
        for sheet_name in file_info['sheets']:
            try:
                # Get sheet dimensions and basic info
                sheet = file_info['workbook'][sheet_name]
                # Instead of sheet.dimensions (not available for ReadOnlyWorksheet), calculate max row/col
                max_row = 0
                max_col = 0
                for row in sheet.iter_rows():
                    max_row += 1
                    if len(row) > max_col:
                        max_col = len(row)
                dimensions = f"Rows: {max_row}, Columns: {max_col}"
                # Read a sample to get column names
                df_sample = pd.read_excel(file_path, sheet_name=sheet_name, nrows=5)
                sheet_info[sheet_name] = {
                    'dimensions': dimensions,
                    'columns': list(df_sample.columns),
                    'column_count': len(df_sample.columns),
                    'sample_rows': len(df_sample)
                }
            except Exception as e:
                sheet_info[sheet_name] = {
                    'error': f"Error reading sheet: {str(e)}"
                }
        
        return sheet_info
    
    def extract_data(self, file_path: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
        """
        Extract data from a specific sheet or all sheets.
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Specific sheet name (if None, returns first sheet)
            
        Returns:
            DataFrame containing the extracted data
        """
        if file_path not in self.loaded_files:
            raise ValueError(f"File {file_path} not loaded. Use load_files() first.")
        
        file_info = self.loaded_files[file_path]
        
        if sheet_name is None:
            sheet_name = file_info['sheets'][0]
        
        if sheet_name not in file_info['sheets']:
            raise ValueError(f"Sheet '{sheet_name}' not found in file {file_path}")
        
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            return df
        except Exception as e:
            raise Exception(f"Error extracting data from sheet '{sheet_name}': {str(e)}")
    
    def preview_data(self, file_path: str, sheet_name: Optional[str] = None, rows: int = 5) -> pd.DataFrame:
        """
        Preview data from a specific sheet.
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Specific sheet name (if None, uses first sheet)
            rows: Number of rows to preview
            
        Returns:
            DataFrame with preview data
        """
        if file_path not in self.loaded_files:
            raise ValueError(f"File {file_path} not loaded. Use load_files() first.")
        
        file_info = self.loaded_files[file_path]
        
        if sheet_name is None:
            sheet_name = file_info['sheets'][0]
        
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=rows)
            return df
        except Exception as e:
            raise Exception(f"Error previewing data from sheet '{sheet_name}': {str(e)}")
    
    def get_all_sheets_data(self, file_path: str) -> Dict[str, pd.DataFrame]:
        """
        Extract data from all sheets in a file.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Dictionary with sheet names as keys and DataFrames as values
        """
        if file_path not in self.loaded_files:
            raise ValueError(f"File {file_path} not loaded. Use load_files() first.")
        
        file_info = self.loaded_files[file_path]
        all_data = {}
        
        for sheet_name in file_info['sheets']:
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                all_data[sheet_name] = df
            except Exception as e:
                print(f"Warning: Could not read sheet '{sheet_name}': {str(e)}")
        
        return all_data
